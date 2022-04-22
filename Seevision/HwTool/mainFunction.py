# coding = utf8
import os
import sys
from time import sleep

import openpyxl
import pandas as pd
import pyautogui

os.path.abspath(".")
"""
    @Project:PyQt5Project
    @File:mainFunction.py
    @Author:十二点前要睡觉
    @Date:2022/4/20 15:58
"""
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QMessageBox, QFileDialog

from PadsElementScratch import Ui_MainWindow

os.path.abspath(".")

# 设置递归限制，防止程序中的循环使得python内存溢出
sys.setrecursionlimit(5000)


def chooseExcelFile():
    global excel_path
    global layer
    excelFile = QFileDialog.getOpenFileName()
    if excelFile:
        excel_path = str(excelFile[0])
        layer = ui.layerChooseCombox.currentText()
        ui.choosedFilePathTextEdit.setText(excel_path)
        ui.runningStatusTextEdit.setText(
            "当前选择的是【{}】层\nExcel文件路径为：【{}】".format(layer, excel_path))


def openSearchBox(name="D15", zoom_times=6):
    pyautogui.press("s")
    pyautogui.press("s")
    pyautogui.typewrite(" {}".format(name))
    pyautogui.press("enter")
    pyautogui.hotkey("ctrl", "w")
    for i in range(zoom_times):
        pyautogui.doubleClick()
    pyautogui.hotkey("ctrl", "w")


def resize_home():
    pyautogui.press("HOME")


def catchFramePicture(name, e_type, e_degree):
    global zoomtimes, GoodX, GoodY
    print("B")
    # C13 _horizontal (0,180)
    # 0402 zoom_times = 6
    e0402_gx_horizontal = screenX * 0.29
    e0402_gy_horizontal = screenY * 0.25
    # vertical (90, 270)  264 * 558 zoom_times = 6
    e0402_gx_vertical = screenX * 0.14
    e0402_gy_vertical = screenY * 0.5

    # C20 _horizontal
    # 0603 zoom_times = 6
    e0603_gx_horizontal = screenX * 0.42
    e0603_gy_horizontal = screenY * 0.28
    # 305 * 824 zoom_times = 6
    e0603_gx_vertical = screenX * 0.17
    e0603_gy_vertical = screenY * 0.76

    # R42 _horizontal
    # 0805 zoom_times = 6
    e0805_gx_horizontal = screenX * 0.64
    e0805_gy_horizontal = screenY * 0.55
    # zoom_times = 5 383*773
    e0805_gx_vertical = screenX * 0.2
    e0805_gy_vertical = screenY * 0.72

    # D17 _horizontal
    # 0603-2 zoom_times = 6
    e0201_gx_horizontal = screenX * 0.22
    e0201_gy_horizontal = screenY * 0.21
    # 229 * 422 zoom_times = 6
    e0201_gx_vertical = screenX * 0.14
    e0201_gy_vertical = screenY * 0.42

    if "0402" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0402_gx_horizontal
            GoodY = e0402_gy_horizontal
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0402_gx_vertical
            GoodY = e0402_gy_vertical
        zoomtimes = 6
    elif "0603" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0603_gx_horizontal
            GoodY = e0603_gy_horizontal
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0603_gx_vertical
            GoodY = e0603_gy_vertical
        zoomtimes = 6
    elif "0805" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0805_gx_horizontal
            GoodY = e0805_gy_horizontal
            zoomtimes = 6
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0805_gx_vertical
            GoodY = e0805_gy_vertical
            zoomtimes = 5
    elif "0603-2" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0201_gx_horizontal
            GoodY = e0201_gy_horizontal
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0201_gx_vertical
            GoodY = e0201_gy_vertical
        zoomtimes = 6
    else:
        return "Skip"
    print("C")
    openSearchBox(name, zoomtimes)
    if not os.path.exists("./screenshot/"):
        os.mkdir("./screenshot/")
    x, y = pyautogui.position()
    imagePath = "./screenshot/{}.jpeg".format(name)
    catch_x, catch_y = x - GoodX / 2, y - GoodY / 2
    catch_w, catch_h = GoodX, GoodY
    pyautogui.screenshot(imagePath, region=(catch_x, catch_y, catch_w, catch_h))
    resize_home()
    return "已截取，待核对\n {}".format(imagePath)


"""
    test form read - excel operate
"""


# 从excel中读取数据并返回（element）

def read_excel_for_page_element(form="./sytj0101/工作簿1.xlsx", sheet_name="Sheet1"):
    """
    通过Pandas模块读取case测试点内容，用于后续遍历case执行测试
    :param form:待读取case Excel文件路径
    :param sheet_name:待读取case Excel文件指定sheet表名
    :return:返回对应case所在行以及对应改行case测试点
    """
    print("从excel中读取数据（测试数据case）并返回（element）")
    df = pd.read_excel(form, sheet_name=sheet_name, index_col="number", engine="openpyxl")
    test_case_list = []
    for i in range(1, df.shape[0] + 1):
        part_type = df.loc[i, "PartDecal"]
        ref_des = df.loc[i, "RefDes"]
        layer = df.loc[i, "Layer"]
        orientation = df.loc[i, "Orient."]
        test_case_list.append([i, part_type, ref_des, layer, orientation])
    return test_case_list


def write_into_excel(form="./sytj0101/工作簿1.xlsx", sheet_name="Sheet1", row=1, column=12, value="PASS"):
    """
    通过openpyxl模块将每一行case的测试结果写入对应每一行的结果列中
    :param form:待写入case Excel文件路径
    :param sheet_name:待写入case Excel文件指定sheet表名
    :param row:待写入case测试结果所在行
    :param column:待写入case测试结果所在列
    :param value:待写入测试结果
    :return:None
    """
    print("将测试结果写入excel表格对应Case的行 - 测试结果处：【{}】".format(value))
    wb = openpyxl.load_workbook(form)
    ws = wb[sheet_name]
    grid_value = ws.cell(row + 1, column).value
    print("grid value is {}".format(grid_value))
    if grid_value is None:
        ws.cell(row + 1, column).value = value
    wb.save(form)


def runButton():
    global screenX, screenY
    if excel_path == "" or layer == "":
        QMessageBox.critical(mainWindow, "错误：", "运行前请选择需要抓取的层以及选择对应的Excel文件", QMessageBox.Ok)
        ui.runningResultTextEdit.setPlainText("文件抓取失败，请检查所需参数以及表格格式是否正常！")
    else:
        QMessageBox.information(mainWindow, "提示：", "运行后，请立即在5s内切换到PADS Layout界面", QMessageBox.Ok)
        sleep(5)
        element_list = read_excel_for_page_element(form=excel_path)
        screenX, screenY = pyautogui.size()
        testSide = layer
        form = excel_path
        for element in element_list:
            e_row = element[0]
            e_type = element[1]
            e_name = element[2]
            e_topOrbottom = element[3]
            e_degree = element[4]
            if e_topOrbottom in testSide:
                result = catchFramePicture(e_name, e_type, e_degree)
                if result == "Skip":
                    pass
                else:
                    write_into_excel(form=form, sheet_name="Sheet1", row=e_row, column=12, value=result)
            else:
                write_into_excel(form=form, sheet_name="Sheet1", row=e_row, column=12, value="")
        print("Done")
        ui.runningResultTextEdit.setPlainText(
            "本次{}层元件内容抓取完成:\n请查看表格：{}\n请核对元件截图：{}".format(layer, form, "在Screenshot文件夹下"))


def ui_connect():
    ui.chooseFileButton.clicked.connect(chooseExcelFile)
    ui.runningButton.clicked.connect(runButton)


if __name__ == '__main__':
    excel_path = ""
    layer = ""

    app = QtWidgets.QApplication(sys.argv)
    mainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(mainWindow)
    icon = QtGui.QIcon()
    icon.addPixmap(QtGui.QPixmap("./seevi_64.ico"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
    mainWindow.setWindowIcon(icon)
    ui_connect()
    mainWindow.show()
    sys.exit(app.exec_())
